import re
import openpyxl
from django.http import HttpResponse
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import time, timedelta
from django.utils import timezone
from django.shortcuts import get_object_or_404
from django.db.models import Q
from rest_framework.decorators import api_view, parser_classes, permission_classes
from rest_framework.permissions import IsAuthenticated, IsAdminUser, AllowAny
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from rest_framework.response import Response
from rest_framework import status

from django.contrib.auth import authenticate
from rest_framework_simplejwt.tokens import RefreshToken

# Import Model
from .models import (
    Guru, Berita, Pengumuman, Presensi, ProfilSekolah, HeroBanner, AlbumGaleri,
    UserProfile, VideoGaleri, FotoAlbum, User, Event,
    Post, PostLike, PostComment, Note, NoteLike
)

# Import Serializer
from .serializers import (
    EventSerializer, GuruSerializer, BeritaSerializer, PengumumanSerializer, PresensiSerializer, ProfilSekolahSerializer, 
    HeroBannerSerializer, AlbumGaleriSerializer, UserProfileSerializer, VideoGaleriSerializer,
    FotoAlbumSerializer, UserSerializer, UserDirectorySerializer, UserMeSerializer,
    PostSerializer, PostCommentSerializer, NoteSerializer, DEFAULT_AVATAR_PATH, safe_get_guru
)


def format_youtube_embed_url(url):
    if not url:
        return url
    if "youtube.com/embed/" in url:
        return url
    pattern = r'(?:https?:\/\/)?(?:www\.)?(?:youtube\.com\/(?:watch\?v=|v\/)|youtu\.be\/)([\w-]{11})'
    match = re.search(pattern, url)
    if match:
        video_id = match.group(1)
        return f"https://www.youtube.com/embed/{video_id}"
    return url


def update_event_status():
    """Memperbarui status event secara otomatis berdasarkan tanggal saat ini."""
    today = timezone.now().date()
    events = Event.objects.filter(status__in=['Akan Datang', 'Berlangsung'])
    for event in events:
        akhir_event = event.tanggal_selesai if event.tanggal_selesai else event.tanggal_mulai
        if akhir_event < today:
            event.status = 'Selesai'
            event.save(update_fields=['status'])
        elif event.tanggal_mulai <= today <= akhir_event and event.status == 'Akan Datang':
            event.status = 'Berlangsung'
            event.save(update_fields=['status'])


# ==========================================
# PUBLIC APIs (Guest)
# ==========================================
@api_view(['GET'])
@permission_classes([AllowAny])
def get_guru(request):
    guru = Guru.objects.all() 
    serializer = GuruSerializer(guru, many=True, context={'request': request}) 
    return Response(serializer.data, status=status.HTTP_200_OK) 

@api_view(['GET'])
@permission_classes([AllowAny])
def get_guru_detail(request, pk):
    guru = get_object_or_404(Guru, pk=pk)
    guru_serializer = GuruSerializer(guru, context={'request': request})
    data = guru_serializer.data

    berita_list = []
    if guru.user:
        berita = Berita.objects.filter(penulis=guru.user).order_by('-tanggal_upload')
        berita_serializer = BeritaSerializer(berita, many=True, context={'request': request})
        berita_list = berita_serializer.data

    data['berita_list'] = berita_list
    return Response(data, status=status.HTTP_200_OK)

@api_view(['GET'])
@permission_classes([AllowAny])
def get_berita(request):
    berita = Berita.objects.all()
    kategori = request.GET.get('kategori', None)
    search_query = request.GET.get('search', None)
    if kategori and kategori != 'Semua':
        berita = berita.filter(kategori=kategori)
    if search_query:
        berita = berita.filter(
            Q(judul__icontains=search_query) | 
            Q(ringkasan__icontains=search_query) | 
            Q(isi__icontains=search_query)
        )
    serializer = BeritaSerializer(berita, many=True, context={'request': request})
    return Response(serializer.data, status=status.HTTP_200_OK)

@api_view(['GET'])
@permission_classes([AllowAny])
def get_berita_detail(request, pk):
    berita = get_object_or_404(Berita, pk=pk)
    serializer = BeritaSerializer(berita, context={'request': request})
    return Response(serializer.data, status=status.HTTP_200_OK)

@api_view(['GET'])
@permission_classes([AllowAny])
def get_pengumuman(request):
    pengumuman = Pengumuman.objects.all().order_by('-tanggal_dibuat') 
    serializer = PengumumanSerializer(pengumuman, many=True, context={'request': request}) 
    return Response(serializer.data, status=status.HTTP_200_OK)

@api_view(['GET'])
@permission_classes([AllowAny])
def get_pengumuman_detail(request, pk):
    pengumuman = get_object_or_404(Pengumuman, pk=pk)
    serializer = PengumumanSerializer(pengumuman, context={'request': request})
    return Response(serializer.data, status=status.HTTP_200_OK)

@api_view(['GET'])
@permission_classes([AllowAny])
def get_profil(request):
    profil = ProfilSekolah.objects.first() 
    if profil:
        serializer = ProfilSekolahSerializer(profil, context={'request': request})
        return Response(serializer.data, status=status.HTTP_200_OK)
    return Response({"message": "Data profil belum diisi"}, status=status.HTTP_404_NOT_FOUND)

@api_view(['GET'])
@permission_classes([AllowAny])
def get_susunan_organisasi(request):
    semua_guru = Guru.objects.all() 
    serializer = GuruSerializer(semua_guru, many=True, context={'request': request})
    return Response(serializer.data, status=status.HTTP_200_OK)

@api_view(['GET'])
@permission_classes([AllowAny])
def get_hero(request):
    banner = HeroBanner.objects.filter(is_active=True)
    serializer = HeroBannerSerializer(banner, many=True, context={'request': request})
    return Response(serializer.data, status=status.HTTP_200_OK)

@api_view(['GET'])
@permission_classes([AllowAny])
def get_galeri(request):
    albums = AlbumGaleri.objects.prefetch_related('fotos').all()
    videos = VideoGaleri.objects.all()
    album_serializer = AlbumGaleriSerializer(albums, many=True, context={'request': request})
    video_serializer = VideoGaleriSerializer(videos, many=True, context={'request': request})
    return Response({
        'albums': album_serializer.data,
        'videos': video_serializer.data
    }, status=status.HTTP_200_OK)

@api_view(['GET'])
@permission_classes([AllowAny])
def get_event(request):
    update_event_status()
    events = Event.objects.all()
    serializer = EventSerializer(events, many=True, context={'request': request})
    return Response(serializer.data, status=status.HTTP_200_OK)

@api_view(['GET'])
@permission_classes([AllowAny])
def get_event_detail(request, pk):
    update_event_status()
    event = get_object_or_404(Event, pk=pk)
    serializer = EventSerializer(event, context={'request': request})
    return Response(serializer.data, status=status.HTTP_200_OK)


# ==========================================
# AUTHENTICATION API (JWT LOGIN, LOGOUT & ME)
# ==========================================
@api_view(['POST'])
@permission_classes([AllowAny])
def login_view(request):
    identifier = str(request.data.get('identifier', '')).strip()
    password = str(request.data.get('password', '')).strip()

    if not identifier or not password:
        return Response({'message': 'Silakan isi Username/NIP dan Kata Sandi.'}, status=status.HTTP_400_BAD_REQUEST)

    user = User.objects.filter(username__iexact=identifier).first()
    if not user:
        guru_obj = Guru.objects.filter(nip=identifier).first()
        if guru_obj and guru_obj.user:
            user = guru_obj.user

    if user is None:
        return Response({'message': 'Kredensial Username/NIP tidak ditemukan.'}, status=status.HTTP_401_UNAUTHORIZED)

    authenticated_user = authenticate(username=user.username, password=password)
    if not authenticated_user:
        return Response({'message': 'Kata sandi yang Anda masukkan salah.'}, status=status.HTTP_401_UNAUTHORIZED)

    if not authenticated_user.is_active:
        return Response({'message': 'Akun Anda sedang tidak aktif. Hubungi Administrator.'}, status=status.HTTP_403_FORBIDDEN)

    roles = []
    if authenticated_user.is_staff or authenticated_user.is_superuser:
        roles.append('admin')

    secondary_role = 'guru'
    if hasattr(authenticated_user, 'guru_profile') and authenticated_user.guru_profile:
        kat = (authenticated_user.guru_profile.kategori or '').lower()
        if 'staf' in kat or 'tu' in kat:
            secondary_role = 'staf'
        else:
            secondary_role = 'guru'
    elif hasattr(authenticated_user, 'profil') and authenticated_user.profil and authenticated_user.profil.role_type:
        secondary_role = authenticated_user.profil.role_type

    if secondary_role not in roles:
        roles.append(secondary_role)

    refresh = RefreshToken.for_user(authenticated_user)
    foto_profil_url = None
    if hasattr(authenticated_user, 'profil') and authenticated_user.profil and authenticated_user.profil.foto_profil:
        foto_profil_url = request.build_absolute_uri(authenticated_user.profil.foto_profil.url)
    elif hasattr(authenticated_user, 'guru_profile') and authenticated_user.guru_profile and authenticated_user.guru_profile.foto:
        foto_profil_url = request.build_absolute_uri(authenticated_user.guru_profile.foto.url)
    else:
        foto_profil_url = request.build_absolute_uri(DEFAULT_AVATAR_PATH)

    nama_lengkap = authenticated_user.get_full_name()
    if not nama_lengkap and hasattr(authenticated_user, 'guru_profile') and authenticated_user.guru_profile:
        nama_lengkap = authenticated_user.guru_profile.nama
    if not nama_lengkap:
        nama_lengkap = authenticated_user.username

    return Response({
        'access': str(refresh.access_token),
        'refresh': str(refresh),
        'user': {
            'id': authenticated_user.id,
            'username': authenticated_user.username,
            'email': authenticated_user.email,
            'nama_lengkap': nama_lengkap,
            'nama': nama_lengkap,
            'roles': roles,
            'foto_profil': foto_profil_url
        }
    }, status=status.HTTP_200_OK)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def logout_view(request):
    try:
        refresh_token = request.data.get("refresh")
        if refresh_token:
            token = RefreshToken(refresh_token)
            token.blacklist()
        return Response({"message": "Berhasil logout dari sistem."}, status=status.HTTP_205_RESET_CONTENT)
    except Exception as e:
        return Response({"message": f"Gagal melangsungkan logout: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def me_view(request):
    serializer = UserMeSerializer(request.user, context={'request': request})
    return Response(serializer.data, status=status.HTTP_200_OK)


# ==========================================
# ADMIN DASHBOARD & MANAGEMENT APIs (FIXED STATS ALBUM)
# ==========================================
@api_view(['GET'])
@permission_classes([IsAuthenticated, IsAdminUser])
def admin_dashboard_stats(request):
    total_guru = Guru.objects.count()
    total_berita = Berita.objects.count()
    total_pengumuman = Pengumuman.objects.count()
    total_event = Event.objects.count()
    total_users = User.objects.count()
    total_album = AlbumGaleri.objects.count()
    total_video = VideoGaleri.objects.count()

    recent_berita = Berita.objects.all().order_by('-tanggal_upload')[:5]
    berita_serializer = BeritaSerializer(recent_berita, many=True, context={'request': request})

    return Response({
        'total_guru': total_guru,
        'total_berita': total_berita,
        'total_pengumuman': total_pengumuman,
        'total_event': total_event,
        'total_users': total_users,
        'total_album': total_album,
        'total_albums': total_album,
        'total_galeri': total_album,
        'total_video': total_video,
        'recent_berita': berita_serializer.data
    }, status=status.HTTP_200_OK)


# ==========================================
# FITUR PRESENSI (SCAN & RIWAYAT + 1 TAHUN RETENSI)
# ==========================================
@api_view(['POST'])
@permission_classes([IsAuthenticated, IsAdminUser])
def admin_scan_presensi(request):
    raw_value = str(request.data.get('uuid', '')).strip()
    if not raw_value:
        return Response({'message': 'Kode QR / UUID tidak valid atau kosong.'}, status=status.HTTP_400_BAD_REQUEST)

    if raw_value.startswith('USER-UUID-'):
        uuid_code = raw_value.replace('USER-UUID-', '').strip()
    else:
        uuid_code = raw_value

    target_user = None

    profil_user = UserProfile.objects.filter(uuid_code=uuid_code).first()
    if profil_user:
        target_user = profil_user.user

    if not target_user:
        target_user = User.objects.filter(username__iexact=raw_value).first()

    if not target_user:
        guru_obj = Guru.objects.filter(nip=raw_value).first()
        if guru_obj and guru_obj.user:
            target_user = guru_obj.user

    if not target_user and raw_value.isdigit():
        try:
            target_user = User.objects.filter(id=int(raw_value)).first()
        except (ValueError, TypeError, OverflowError):
            pass

    if not target_user:
        return Response(
            {'message': f'Pengguna dengan kredensial "{raw_value}" tidak ditemukan di sistem!'},
            status=status.HTTP_404_NOT_FOUND
        )

    satu_tahun_lalu = timezone.now() - timedelta(days=365)
    Presensi.objects.filter(waktu_scan__lt=satu_tahun_lalu).delete()

    guru_profile = safe_get_guru(target_user)
    nama_lengkap = None
    if guru_profile and guru_profile.nama:
        nama_lengkap = guru_profile.nama
    if not nama_lengkap:
        nama_lengkap = target_user.get_full_name() or target_user.username

    now = timezone.localtime(timezone.now())
    hari_ini = now.date()

    presensi_hari_ini = Presensi.objects.filter(user=target_user, waktu_scan__date=hari_ini).first()

    peran_str = 'Siswa'
    if target_user.is_staff or target_user.is_superuser:
        peran_str = 'Guru/Admin'
    elif guru_profile:
        peran_str = 'Guru'

    if not presensi_hari_ini:
        current_time = now.time()
        batas_toleransi = time(7, 1, 59)
        status_kehadiran = 'Hadir Tepat Waktu' if current_time <= batas_toleransi else 'Terlambat'

        presensi_baru = Presensi.objects.create(
            user=target_user,
            status=status_kehadiran,
            peran=peran_str
        )

        serializer = PresensiSerializer(presensi_baru, context={'request': request})
        waktu_str = now.strftime('%H:%M:%S')

        return Response({
            'type': 'masuk',
            'message': f'BERHASIL MASUK: {nama_lengkap} ({status_kehadiran}) jam {waktu_str}',
            'nama_pengguna': nama_lengkap,
            'status_kehadiran': status_kehadiran,
            'waktu_scan': waktu_str,
            'data': serializer.data
        }, status=status.HTTP_201_CREATED)

    else:
        if presensi_hari_ini.waktu_pulang:
            return Response({
                'message': f'{nama_lengkap} sudah melakukan presensi MASUK dan PULANG hari ini!'
            }, status=status.HTTP_400_BAD_REQUEST)

        waktu_masuk_local = timezone.localtime(presensi_hari_ini.waktu_scan)
        selisih_detik = (now - waktu_masuk_local).total_seconds()

        if selisih_detik < 1800:
            menit_sisa = int((1800 - selisih_detik) // 60) + 1
            return Response({
                'message': f'Gagal Scan Pulang! Minimal 30 menit setelah scan masuk baru bisa absen pulang. Silakan tunggu sekitar {menit_sisa} menit lagi.'
            }, status=status.HTTP_400_BAD_REQUEST)

        presensi_hari_ini.waktu_pulang = now
        presensi_hari_ini.save()

        serializer = PresensiSerializer(presensi_hari_ini, context={'request': request})
        waktu_str = now.strftime('%H:%M:%S')

        return Response({
            'type': 'pulang',
            'message': f'BERHASIL PULANG: {nama_lengkap} absen pulang jam {waktu_str}',
            'nama_pengguna': nama_lengkap,
            'status_kehadiran': presensi_hari_ini.status,
            'waktu_scan': presensi_hari_ini.waktu_scan.strftime('%H:%M:%S'),
            'waktu_pulang': waktu_str,
            'data': serializer.data
        }, status=status.HTTP_200_OK)


@api_view(['GET'])
@permission_classes([IsAuthenticated, IsAdminUser])
def admin_export_presensi_excel(request):
    search = request.GET.get('search', '').strip()
    peran_filter = request.GET.get('peran', '').strip()

    riwayat = Presensi.objects.all().order_by('-waktu_scan')
    
    if search:
        riwayat = riwayat.filter(
            Q(user__username__icontains=search) |
            Q(user__first_name__icontains=search) |
            Q(user__last_name__icontains=search) |
            Q(user__guru_profile__nama__icontains=search) |
            Q(user__guru_profile__nip__icontains=search)
        )
        
    if peran_filter and peran_filter != 'Semua Peran':
        riwayat = riwayat.filter(peran__icontains=peran_filter)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Riwayat Presensi SD YWKA"

    font_title = Font(name='Calibri', size=14, bold=True, color="1E293B")
    font_header = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    font_data = Font(name='Calibri', size=11)
    
    fill_header = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    
    border_thin = Side(border_style="thin", color="CBD5E1")
    box_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)

    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")

    ws.merge_cells('A1:J1')
    ws['A1'] = "REKAPITULASI PRESENSI KEHADIRAN PEGAWAI & SISWA"
    ws['A1'].font = font_title
    ws['A1'].alignment = align_center

    ws.merge_cells('A2:J2')
    waktu_cetak = timezone.localtime(timezone.now()).strftime('%d-%m-%Y %H:%M:%S')
    ws['A2'] = f"Waktu Cetak System: {waktu_cetak} WIB"
    ws['A2'].font = Font(name='Calibri', size=10, italic=True, color="64748B")
    ws['A2'].alignment = align_center

    headers = [
        "No", "Tanggal", "Nama Pengguna", "Username", 
        "NIP / NIY", "Jabatan / Sub Info", "Peran", 
        "Jam Masuk", "Jam Pulang", "Status Kehadiran"
    ]
    
    ws.append([])
    
    ws.row_dimensions[4].height = 25
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = box_border

    row_num = 5
    for idx, p in enumerate(riwayat, 1):
        guru = safe_get_guru(p.user)
        
        nama_pengguna = guru.nama if (guru and guru.nama) else (p.user.get_full_name() or p.user.username)
        username = p.user.username
        nip = guru.nip if (guru and guru.nip) else "-"
        jabatan = guru.jabatan if (guru and guru.jabatan) else ("Admin/Staf" if p.user.is_staff else "Siswa / Pengguna")
        
        waktu_masuk_local = timezone.localtime(p.waktu_scan)
        tanggal_str = waktu_masuk_local.strftime('%d-%m-%Y')
        jam_masuk_str = waktu_masuk_local.strftime('%H:%M:%S WIB')
        
        if p.waktu_pulang:
            jam_pulang_str = timezone.localtime(p.waktu_pulang).strftime('%H:%M:%S WIB')
        else:
            jam_pulang_str = "-"

        row_data = [
            idx,
            tanggal_str,
            nama_pengguna,
            username,
            nip,
            jabatan,
            p.peran,
            jam_masuk_str,
            jam_pulang_str,
            p.status
        ]

        ws.append(row_data)
        ws.row_dimensions[row_num].height = 20

        is_even = (row_num % 2 == 0)
        for col_idx in range(1, len(row_data) + 1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.font = font_data
            cell.border = box_border

            if is_even:
                cell.fill = fill_zebra

            if col_idx in [1, 2, 7, 8, 9, 10]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left

            if col_idx == 10:
                if p.status == 'Hadir Tepat Waktu':
                    cell.font = Font(name='Calibri', size=11, bold=True, color="15803D")
                else:
                    cell.font = Font(name='Calibri', size=11, bold=True, color="B45309")

        row_num += 1

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row < 4:
                continue
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    file_name = f"Riwayat_Presensi_{timezone.localtime(timezone.now()).strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{file_name}"'

    wb.save(response)
    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated, IsAdminUser])
def admin_riwayat_presensi(request):
    search = request.GET.get('search', '').strip()
    peran_filter = request.GET.get('peran', '').strip()

    riwayat = Presensi.objects.all().order_by('-waktu_scan')
    if search:
        riwayat = riwayat.filter(
            Q(user__username__icontains=search) |
            Q(user__first_name__icontains=search) |
            Q(user__last_name__icontains=search) |
            Q(user__guru_profile__nama__icontains=search) |
            Q(user__guru_profile__nip__icontains=search)
        )
    if peran_filter and peran_filter != 'Semua Peran':
        riwayat = riwayat.filter(peran__icontains=peran_filter)

    serializer = PresensiSerializer(riwayat, many=True, context={'request': request})
    return Response(serializer.data, status=status.HTTP_200_OK)


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated, IsAdminUser])
@parser_classes([MultiPartParser, FormParser, JSONParser])
def admin_berita_list_create(request):
    if request.method == 'GET':
        berita = Berita.objects.all()
        serializer = BeritaSerializer(berita, many=True, context={'request': request})
        return Response(serializer.data, status=status.HTTP_200_OK)
    
    elif request.method == 'POST':
        serializer = BeritaSerializer(data=request.data, context={'request': request})
        if serializer.is_valid():
            serializer.save(penulis=request.user)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated, IsAdminUser])
@parser_classes([MultiPartParser, FormParser, JSONParser])
def admin_berita_detail(request, pk):
    berita = get_object_or_404(Berita, pk=pk)

    if request.method == 'GET':
        serializer = BeritaSerializer(berita, context={'request': request})
        return Response(serializer.data, status=status.HTTP_200_OK)

    elif request.method == 'PUT':
        serializer = BeritaSerializer(berita, data=request.data, partial=True, context={'request': request})
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    elif request.method == 'DELETE':
        berita.delete()
        return Response({"message": "Berita berhasil dihapus."}, status=status.HTTP_204_NO_CONTENT)


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated, IsAdminUser])
@parser_classes([MultiPartParser, FormParser, JSONParser])
def admin_pengumuman_list_create(request):
    if request.method == 'GET':
        pengumuman = Pengumuman.objects.all()
        serializer = PengumumanSerializer(pengumuman, many=True, context={'request': request})
        return Response(serializer.data, status=status.HTTP_200_OK)

    elif request.method == 'POST':
        serializer = PengumumanSerializer(data=request.data, context={'request': request})
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated, IsAdminUser])
@parser_classes([MultiPartParser, FormParser, JSONParser])
def admin_pengumuman_detail(request, pk):
    pengumuman = get_object_or_404(Pengumuman, pk=pk)

    if request.method == 'GET':
        serializer = PengumumanSerializer(pengumuman, context={'request': request})
        return Response(serializer.data, status=status.HTTP_200_OK)

    elif request.method == 'PUT':
        serializer = PengumumanSerializer(pengumuman, data=request.data, partial=True, context={'request': request})
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    elif request.method == 'DELETE':
        pengumuman.delete()
        return Response({"message": "Pengumuman berhasil dihapus."}, status=status.HTTP_204_NO_CONTENT)


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated, IsAdminUser])
@parser_classes([MultiPartParser, FormParser, JSONParser])
def admin_guru_list_create(request):
    if request.method == 'GET':
        guru = Guru.objects.all()
        serializer = GuruSerializer(guru, many=True, context={'request': request})
        return Response(serializer.data, status=status.HTTP_200_OK)

    elif request.method == 'POST':
        serializer = GuruSerializer(data=request.data, context={'request': request})
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated, IsAdminUser])
@parser_classes([MultiPartParser, FormParser, JSONParser])
def admin_guru_detail(request, pk):
    guru = get_object_or_404(Guru, pk=pk)

    if request.method == 'GET':
        serializer = GuruSerializer(guru, context={'request': request})
        return Response(serializer.data, status=status.HTTP_200_OK)

    elif request.method == 'PUT':
        serializer = GuruSerializer(guru, data=request.data, partial=True, context={'request': request})
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    elif request.method == 'DELETE':
        guru.delete()
        return Response({"message": "Data Guru/SDM berhasil dihapus."}, status=status.HTTP_204_NO_CONTENT)


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated, IsAdminUser])
@parser_classes([MultiPartParser, FormParser, JSONParser])
def admin_album_list_create(request):
    if request.method == 'GET':
        albums = AlbumGaleri.objects.prefetch_related('fotos').all()
        serializer = AlbumGaleriSerializer(albums, many=True, context={'request': request})
        return Response(serializer.data, status=status.HTTP_200_OK)

    elif request.method == 'POST':
        serializer = AlbumGaleriSerializer(data=request.data, context={'request': request})
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated, IsAdminUser])
@parser_classes([MultiPartParser, FormParser, JSONParser])
def admin_album_detail(request, pk):
    album = get_object_or_404(AlbumGaleri.objects.prefetch_related('fotos'), pk=pk)

    if request.method == 'GET':
        serializer = AlbumGaleriSerializer(album, context={'request': request})
        return Response(serializer.data, status=status.HTTP_200_OK)

    elif request.method == 'PUT':
        serializer = AlbumGaleriSerializer(album, data=request.data, partial=True, context={'request': request})
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    elif request.method == 'DELETE':
        album.delete()
        return Response({"message": "Album galeri berhasil dihapus."}, status=status.HTTP_204_NO_CONTENT)


@api_view(['POST'])
@permission_classes([IsAuthenticated, IsAdminUser])
@parser_classes([MultiPartParser, FormParser])
def admin_foto_upload(request, album_pk):
    album = get_object_or_404(AlbumGaleri, pk=album_pk)
    
    uploaded_files = []
    for key in ['gambar', 'foto', 'file', 'fotos', 'files']:
        if key in request.FILES:
            uploaded_files.extend(request.FILES.getlist(key))

    if not uploaded_files:
        serializer = FotoAlbumSerializer(data=request.data, context={'request': request})
        if serializer.is_valid():
            serializer.save(album=album)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(
            {'message': 'Gagal mengunggah foto. Pastikan memilih file gambar yang valid dengan key (gambar/foto/file).'}, 
            status=status.HTTP_400_BAD_REQUEST
        )

    created_fotos = []
    keterangan = request.data.get('keterangan', '')

    for file_obj in uploaded_files:
        foto_item = FotoAlbum.objects.create(
            album=album,
            gambar=file_obj,
            keterangan=keterangan
        )
        created_fotos.append(foto_item)

    serializer = FotoAlbumSerializer(created_fotos, many=True, context={'request': request})
    return Response(serializer.data, status=status.HTTP_201_CREATED)


@api_view(['DELETE'])
@permission_classes([IsAuthenticated, IsAdminUser])
def admin_foto_delete(request, pk):
    foto = get_object_or_404(FotoAlbum, pk=pk)
    foto.delete()
    return Response({"message": "Foto item berhasil dihapus."}, status=status.HTTP_204_NO_CONTENT)


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated, IsAdminUser])
def admin_video_list_create(request):
    if request.method == 'GET':
        videos = VideoGaleri.objects.all()
        serializer = VideoGaleriSerializer(videos, many=True, context={'request': request})
        return Response(serializer.data, status=status.HTTP_200_OK)

    elif request.method == 'POST':
        data = request.data.copy()
        if 'embed_url' in data:
            data['embed_url'] = format_youtube_embed_url(data['embed_url'])

        serializer = VideoGaleriSerializer(data=data, context={'request': request})
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated, IsAdminUser])
def admin_video_detail(request, pk):
    video = get_object_or_404(VideoGaleri, pk=pk)

    if request.method == 'GET':
        serializer = VideoGaleriSerializer(video, context={'request': request})
        return Response(serializer.data, status=status.HTTP_200_OK)

    elif request.method == 'PUT':
        data = request.data.copy()
        if 'embed_url' in data:
            data['embed_url'] = format_youtube_embed_url(data['embed_url'])

        serializer = VideoGaleriSerializer(video, data=data, partial=True, context={'request': request})
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    elif request.method == 'DELETE':
        video.delete()
        return Response({"message": "Video galeri berhasil dihapus."}, status=status.HTTP_204_NO_CONTENT)


@api_view(['GET', 'PUT'])
@permission_classes([IsAuthenticated, IsAdminUser])
def admin_profil_detail_update(request):
    profil, _ = ProfilSekolah.objects.get_or_create(id=1)

    if request.method == 'GET':
        serializer = ProfilSekolahSerializer(profil, context={'request': request})
        return Response(serializer.data, status=status.HTTP_200_OK)

    elif request.method == 'PUT':
        serializer = ProfilSekolahSerializer(profil, data=request.data, partial=True, context={'request': request})
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated, IsAdminUser])
@parser_classes([MultiPartParser, FormParser, JSONParser])
def admin_hero_banner_list_create(request):
    if request.method == 'GET':
        banners = HeroBanner.objects.all()
        serializer = HeroBannerSerializer(banners, many=True, context={'request': request})
        return Response(serializer.data, status=status.HTTP_200_OK)

    elif request.method == 'POST':
        serializer = HeroBannerSerializer(data=request.data, context={'request': request})
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated, IsAdminUser])
@parser_classes([MultiPartParser, FormParser, JSONParser])
def admin_hero_banner_detail(request, pk):
    banner = get_object_or_404(HeroBanner, pk=pk)

    if request.method == 'GET':
        serializer = HeroBannerSerializer(banner, context={'request': request})
        return Response(serializer.data, status=status.HTTP_200_OK)

    elif request.method == 'PUT':
        serializer = HeroBannerSerializer(banner, data=request.data, partial=True, context={'request': request})
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    elif request.method == 'DELETE':
        banner.delete()
        return Response({"message": "Banner hero berhasil dihapus."}, status=status.HTTP_204_NO_CONTENT)


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated, IsAdminUser])
@parser_classes([MultiPartParser, FormParser, JSONParser])
def admin_user_list_create(request):
    if request.method == 'GET':
        users = User.objects.all().order_by('-id')
        serializer = UserSerializer(users, many=True, context={'request': request})
        return Response(serializer.data, status=status.HTTP_200_OK)

    elif request.method == 'POST':
        serializer = UserSerializer(data=request.data, context={'request': request})
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated, IsAdminUser])
@parser_classes([MultiPartParser, FormParser, JSONParser])
def admin_user_detail(request, pk):
    user = get_object_or_404(User, pk=pk)

    if request.method == 'GET':
        serializer = UserSerializer(user, context={'request': request})
        return Response(serializer.data, status=status.HTTP_200_OK)

    elif request.method == 'PUT':
        serializer = UserSerializer(user, data=request.data, partial=True, context={'request': request})
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    elif request.method == 'DELETE':
        user.delete()
        return Response({"message": "User berhasil dihapus."}, status=status.HTTP_204_NO_CONTENT)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def user_list(request):
    users = User.objects.filter(is_active=True).order_by('first_name', 'username')
    serializer = UserDirectorySerializer(users, many=True, context={'request': request})
    return Response(serializer.data, status=status.HTTP_200_OK)


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated, IsAdminUser])
@parser_classes([MultiPartParser, FormParser, JSONParser])
def admin_event_list_create(request):
    update_event_status()
    if request.method == 'GET':
        events = Event.objects.all()
        serializer = EventSerializer(events, many=True, context={'request': request})
        return Response(serializer.data, status=status.HTTP_200_OK)

    elif request.method == 'POST':
        serializer = EventSerializer(data=request.data, context={'request': request})
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated, IsAdminUser])
@parser_classes([MultiPartParser, FormParser, JSONParser])
def admin_event_detail(request, pk):
    update_event_status()
    event = get_object_or_404(Event, pk=pk)

    if request.method == 'GET':
        serializer = EventSerializer(event, context={'request': request})
        return Response(serializer.data, status=status.HTTP_200_OK)

    elif request.method == 'PUT':
        serializer = EventSerializer(event, data=request.data, partial=True, context={'request': request})
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    elif request.method == 'DELETE':
        event.delete()
        return Response({"message": "Event berhasil dihapus."}, status=status.HTTP_204_NO_CONTENT)


# ==========================================
# POSTINGAN, LIKE & KOMENTAR (USER PORTAL)
# ==========================================
@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
@parser_classes([MultiPartParser, FormParser, JSONParser])
def user_post_list_create(request):
    if request.method == 'GET':
        user_id = request.query_params.get('user_id') or request.GET.get('user_id')
        if user_id:
            posts = Post.objects.filter(user_id=user_id)
        else:
            posts = Post.objects.all()
            
        serializer = PostSerializer(posts, many=True, context={'request': request})
        return Response(serializer.data, status=status.HTTP_200_OK)

    elif request.method == 'POST':
        serializer = PostSerializer(data=request.data, context={'request': request})
        if serializer.is_valid():
            serializer.save(user=request.user)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def user_my_posts(request):
    posts = Post.objects.filter(user=request.user)
    serializer = PostSerializer(posts, many=True, context={'request': request})
    return Response(serializer.data, status=status.HTTP_200_OK)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def user_post_like_toggle(request, pk):
    post = get_object_or_404(Post, pk=pk)
    like_obj = PostLike.objects.filter(post=post, user=request.user).first()
    if like_obj:
        like_obj.delete()
        is_liked = False
    else:
        PostLike.objects.create(post=post, user=request.user)
        is_liked = True

    return Response({
        'is_liked': is_liked,
        'likes_count': post.likes.count()
    }, status=status.HTTP_200_OK)


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def user_post_comment_list_create(request, pk):
    post = get_object_or_404(Post, pk=pk)

    if request.method == 'GET':
        comments = post.comments.all()
        serializer = PostCommentSerializer(comments, many=True, context={'request': request})
        return Response(serializer.data, status=status.HTTP_200_OK)

    elif request.method == 'POST':
        serializer = PostCommentSerializer(data=request.data, context={'request': request})
        if serializer.is_valid():
            serializer.save(user=request.user, post=post)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def user_post_delete(request, pk):
    post = get_object_or_404(Post, pk=pk)
    if post.user != request.user and not (request.user.is_staff or request.user.is_superuser):
        return Response({"message": "Anda tidak memiliki izin untuk menghapus postingan ini."}, status=status.HTTP_403_FORBIDDEN)
    post.delete()
    return Response({"message": "Postingan berhasil dihapus."}, status=status.HTTP_204_NO_CONTENT)


# ==========================================
# STATUS / NOTE TEKS (HILANG 12 JAM)
# ==========================================
@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def user_note_list_create(request):
    if request.method == 'GET':
        time_threshold = timezone.now() - timedelta(hours=12)
        notes = Note.objects.filter(created_at__gte=time_threshold)
        serializer = NoteSerializer(notes, many=True, context={'request': request})
        return Response(serializer.data, status=status.HTTP_200_OK)

    elif request.method == 'POST':
        serializer = NoteSerializer(data=request.data, context={'request': request})
        if serializer.is_valid():
            serializer.save(user=request.user)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def user_note_like_toggle(request, pk):
    note = get_object_or_404(Note, pk=pk)
    like_obj = NoteLike.objects.filter(note=note, user=request.user).first()
    if like_obj:
        like_obj.delete()
        is_liked = False
    else:
        NoteLike.objects.create(note=note, user=request.user)
        is_liked = True

    return Response({
        'is_liked': is_liked,
        'likes_count': note.likes.count()
    }, status=status.HTTP_200_OK)


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def user_note_delete(request, pk):
    note = get_object_or_404(Note, pk=pk)
    if note.user != request.user and not (request.user.is_staff or request.user.is_superuser):
        return Response({"message": "Anda tidak memiliki izin untuk menghapus status ini."}, status=status.HTTP_403_FORBIDDEN)
    note.delete()
    return Response({"message": "Status berhasil dihapus."}, status=status.HTTP_204_NO_CONTENT)